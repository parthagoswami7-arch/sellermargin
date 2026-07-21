"""Admin exports: Sales CSV + GSTR-1 filing Excel."""
from __future__ import annotations
import csv
import io
from datetime import datetime, timezone
from typing import Iterable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _fmt_date(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except Exception:
            return dt[:10]
    return dt.strftime("%d-%b-%Y")


def _dd_mm_yyyy(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except Exception:
            return dt[:10]
    return dt.strftime("%d-%m-%Y")


def build_sales_csv(orders: Iterable[dict], seller: dict) -> bytes:
    """Flat sales CSV — every paid order, one row. Includes GST split."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "Invoice No", "Invoice Date", "Order ID", "CF Payment ID",
        "Buyer Name", "Buyer Email", "Buyer GSTIN", "Buyer State", "Place of Supply",
        "Plan", "Plan Duration (days)", "Reports Added",
        "Taxable Value (Rs.)", "CGST (Rs.)", "SGST (Rs.)", "IGST (Rs.)",
        "Total Tax (Rs.)", "Grand Total (Rs.)",
        "GST Rate %", "SAC Code",
    ])
    for o in orders:
        gst = o.get("gst") or {}
        w.writerow([
            o.get("invoice_no", ""),
            _fmt_date(o.get("invoice_generated_at") or o.get("created_at")),
            o.get("order_id", ""),
            o.get("cf_order_id", ""),
            o.get("buyer_name") or o.get("user_email", ""),
            o.get("user_email", ""),
            o.get("buyer_gstin") or "Unregistered",
            o.get("buyer_state") or seller.get("state", ""),
            o.get("buyer_state") or seller.get("state", ""),
            o.get("plan", ""),
            _plan_days(o),
            _plan_reports(o),
            f"{float(o.get('base_amount') or 0):.2f}",
            f"{float(gst.get('cgst') or 0):.2f}",
            f"{float(gst.get('sgst') or 0):.2f}",
            f"{float(gst.get('igst') or 0):.2f}",
            f"{float(gst.get('total_tax') or 0):.2f}",
            f"{float(o.get('amount') or 0):.2f}",
            "18",
            seller.get("sac_code", "998314"),
        ])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel-friendly UTF-8


def _plan_days(o: dict) -> int:
    return {"trial_10": 7, "annual": 365, "topup_5": 0}.get(o.get("plan"), 0)


def _plan_reports(o: dict) -> int:
    return {"trial_10": 1, "annual": 12, "topup_5": 5}.get(o.get("plan"), 0)


def build_gstr1_excel(orders: Iterable[dict], seller: dict, period_month: int, period_year: int) -> bytes:
    """GSTR-1 filing workbook — separate sheets for B2B (registered buyers) and B2CS
    (unregistered/aggregated). Matches GSTN offline utility column structure so the
    filed CA/accountant can copy-paste into the government's template."""

    wb = Workbook()

    # Cover sheet
    cover = wb.active
    cover.title = "Summary"
    cover["A1"] = f"GSTR-1 Export · {_month_name(period_month)} {period_year}"
    cover["A1"].font = Font(bold=True, size=14, color="044535")
    cover["A3"] = "Seller:"
    cover["B3"] = seller.get("business_name", "")
    cover["A4"] = "GSTIN:"
    cover["B4"] = seller.get("gstin", "")
    cover["A5"] = "State:"
    cover["B5"] = f"{seller.get('state', '')} ({seller.get('state_code', '')})"
    cover["A6"] = "Return Period:"
    cover["B6"] = f"{period_month:02d}{period_year}"
    cover["A7"] = "Generated on:"
    cover["B7"] = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")
    for c in ("A3", "A4", "A5", "A6", "A7"):
        cover[c].font = Font(bold=True)
    cover.column_dimensions["A"].width = 18
    cover.column_dimensions["B"].width = 60

    # Filter to orders paid inside the requested month
    period_orders = []
    for o in orders:
        d = o.get("invoice_generated_at") or o.get("created_at")
        if isinstance(d, str):
            try:
                d = datetime.fromisoformat(d)
            except Exception:
                continue
        if not d:
            continue
        if d.month == period_month and d.year == period_year:
            period_orders.append(o)

    # === Sheet: b2b (registered buyers with GSTIN) ===
    ws_b2b = wb.create_sheet("b2b")
    b2b_headers = [
        "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice date",
        "Invoice Value", "Place Of Supply", "Reverse Charge",
        "Applicable % of Tax Rate", "Invoice Type", "E-Commerce GSTIN",
        "Rate", "Taxable Value", "Cess Amount",
    ]
    _write_header_row(ws_b2b, b2b_headers)

    b2b_orders = [o for o in period_orders if (o.get("buyer_gstin") or "").strip()]
    for row_i, o in enumerate(b2b_orders, start=2):
        gst = o.get("gst") or {}
        pos = _pos_code(o.get("buyer_state") or seller.get("state", ""), seller)
        ws_b2b.cell(row_i, 1, o.get("buyer_gstin", "").strip().upper())
        ws_b2b.cell(row_i, 2, o.get("buyer_name") or "")
        ws_b2b.cell(row_i, 3, o.get("invoice_no") or "")
        ws_b2b.cell(row_i, 4, _dd_mm_yyyy(o.get("invoice_generated_at") or o.get("created_at")))
        ws_b2b.cell(row_i, 5, round(float(o.get("amount") or 0), 2))
        ws_b2b.cell(row_i, 6, pos)
        ws_b2b.cell(row_i, 7, "N")
        ws_b2b.cell(row_i, 8, "")
        ws_b2b.cell(row_i, 9, "Regular B2B")
        ws_b2b.cell(row_i, 10, "")
        ws_b2b.cell(row_i, 11, 18)
        ws_b2b.cell(row_i, 12, round(float(o.get("base_amount") or 0), 2))
        ws_b2b.cell(row_i, 13, 0)

    _autosize(ws_b2b, b2b_headers)

    # === Sheet: b2cs (unregistered — aggregated by state × rate) ===
    ws_b2cs = wb.create_sheet("b2cs")
    b2cs_headers = [
        "Type", "Place Of Supply", "Applicable % of Tax Rate", "Rate",
        "Taxable Value", "Cess Amount", "E-Commerce GSTIN",
    ]
    _write_header_row(ws_b2cs, b2cs_headers)

    # Aggregate: state -> {"taxable": sum, "invoice_value": sum}
    b2cs_agg: dict[str, dict[str, float]] = {}
    for o in period_orders:
        if (o.get("buyer_gstin") or "").strip():
            continue
        state = o.get("buyer_state") or seller.get("state", "")
        pos = _pos_code(state, seller)
        bucket = b2cs_agg.setdefault(pos, {"taxable": 0.0, "invoice_value": 0.0})
        bucket["taxable"] += float(o.get("base_amount") or 0)
        bucket["invoice_value"] += float(o.get("amount") or 0)

    row_i = 2
    for pos, agg in sorted(b2cs_agg.items()):
        ws_b2cs.cell(row_i, 1, "OE")
        ws_b2cs.cell(row_i, 2, pos)
        ws_b2cs.cell(row_i, 3, "")
        ws_b2cs.cell(row_i, 4, 18)
        ws_b2cs.cell(row_i, 5, round(agg["taxable"], 2))
        ws_b2cs.cell(row_i, 6, 0)
        ws_b2cs.cell(row_i, 7, "")
        row_i += 1
    _autosize(ws_b2cs, b2cs_headers)

    # === Summary totals on cover ===
    total_b2b_tax = sum(float(o.get("base_amount") or 0) for o in b2b_orders)
    total_b2b_gst = sum(float((o.get("gst") or {}).get("total_tax") or 0) for o in b2b_orders)
    total_b2cs_tax = sum(a["taxable"] for a in b2cs_agg.values())
    total_b2cs_gst = round(total_b2cs_tax * 0.18, 2)

    cover["A10"] = "Totals for this period"
    cover["A10"].font = Font(bold=True, color="044535")
    cover["A11"] = "B2B taxable value:"
    cover["B11"] = round(total_b2b_tax, 2)
    cover["A12"] = "B2B GST collected:"
    cover["B12"] = round(total_b2b_gst, 2)
    cover["A13"] = "B2CS taxable value:"
    cover["B13"] = round(total_b2cs_tax, 2)
    cover["A14"] = "B2CS GST collected:"
    cover["B14"] = round(total_b2cs_gst, 2)
    cover["A15"] = "Total taxable:"
    cover["B15"] = round(total_b2b_tax + total_b2cs_tax, 2)
    cover["A15"].font = Font(bold=True)
    cover["B15"].font = Font(bold=True)
    cover["A16"] = "Total GST payable:"
    cover["B16"] = round(total_b2b_gst + total_b2cs_gst, 2)
    cover["A16"].font = Font(bold=True, color="044535")
    cover["B16"].font = Font(bold=True, color="044535")

    cover["A18"] = ("Instructions: Open the GSTR-1 Offline Utility from gst.gov.in, "
                    "then copy the rows from the b2b and b2cs sheets into the matching tabs of the utility "
                    "template. Column order matches the government template exactly.")
    cover["A18"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[18].height = 40
    cover.merge_cells("A18:D18")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header_row(ws, headers):
    fill = PatternFill("solid", fgColor="044535")
    thin = Side(style="thin", color="D8DAD5")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border


def _autosize(ws, headers):
    for i, h in enumerate(headers, 1):
        # Rough: max of header length or 14, capped at 30
        w = max(len(h) + 2, 14)
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = min(w, 30)


def _pos_code(state_name: str, seller: dict) -> str:
    """Return 'CC-State Name' as GSTN expects, e.g. '27-Maharashtra'."""
    from invoice import STATE_CODES
    code = STATE_CODES.get(state_name)
    if not code:
        code = seller.get("state_code", "")
    return f"{code}-{state_name}" if code and state_name else (state_name or "")


def _month_name(m: int) -> str:
    return ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug",
            "Sep", "Oct", "Nov", "Dec"][m - 1]
