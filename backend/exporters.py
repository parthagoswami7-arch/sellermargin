"""Excel and PDF exporters for P&L reports."""
from __future__ import annotations
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

EMERALD = "044535"
GOLD    = "F4B223"
ORANGE  = "FFE9CC"


def _month_name(m: int) -> str:
    return ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][m - 1]


def export_excel(report: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Summary"

    s = report["summary"]
    rows_data = report["rows"]

    thin = Side(style="thin", color="D8DAD5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"AMAZON MONTHLY P&L — {_month_name(s['target_month']).upper()} {s['target_year']}"
    ws["A1"].font = Font(bold=True, size=16, color=EMERALD)
    ws.merge_cells("A1:D1")

    summary_rows = [
        ("Settlement (total payment received)", s["settlement"]),
        ("Reimbursement (FBA adjustments)", s["reimbursement"]),
        ("TOTAL RECEIVED", s["total_received"]),
        ("Cost price (COGS total)", s["cogs"]),
        ("Inbound (FBA Inbound Pickup Service)", s["inbound_fee"]),
        ("Storage fee (posts 7th of following month)", s["storage_fee"]),
        ("Outbound / Removal fee", s["removal_fee"]),
        ("Ad spend (Sponsored Products)", s["ad_spend"]),
        ("TOTAL DEDUCTION", s["total_deduction"]),
        ("FINAL PROFIT", s["final_profit"]),
        ("", ""),
        ("ACOS % (Ad spend / Item price total)", f"{s['acos_pct']:.2f}%"),
        ("Profit % (Final profit / Item price total)", f"{s['profit_pct']:.2f}%"),
        ("Profit % on Cost price", f"{s['profit_pct_on_cogs']:.2f}%"),
        ("Customer Return %", f"{s.get('customer_return_pct', s.get('return_pct', 0)):.2f}%"),
        ("RTO %", f"{s.get('rto_pct', 0):.2f}%"),
    ]

    r = 3
    for label, val in summary_rows:
        ws.cell(r, 1, label).font = Font(bold=(label in ("TOTAL RECEIVED", "TOTAL DEDUCTION", "FINAL PROFIT")))
        c = ws.cell(r, 2, val)
        c.alignment = Alignment(horizontal="right")
        if label == "FINAL PROFIT":
            c.font = Font(bold=True, size=14, color=EMERALD)
        r += 1

    # GST reconciliation note
    r += 1
    note_cell = ws.cell(r, 1, "Note: Net profit = Gross profit (Final Profit above) − Net GST payable")
    note_cell.font = Font(italic=True, color=EMERALD)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    note2 = ws.cell(r, 1, "GST liability from GSTR filings is not subtracted in this sheet — please deduct separately.")
    note2.font = Font(italic=True, color="5C5F5A", size=10)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 20

    # Second sheet: rows
    ws2 = wb.create_sheet("Rows")
    headers = ["order-id", "sku", "quantity", "item-price", "payment",
               "cost-price-unit", "total-cost", "return-reason",
               "product-condition", "order-status"]
    for i, h in enumerate(headers, 1):
        c = ws2.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=EMERALD)
        c.alignment = Alignment(horizontal="left")

    orange_fill = PatternFill("solid", fgColor=ORANGE)
    for idx, row in enumerate(rows_data, start=2):
        eff_cost = row.get("cost_price_unit_override") if row.get("cost_price_unit_override") is not None else (row.get("cost_price_unit") or 0)
        total_cost = float(row["quantity"]) * float(eff_cost or 0)
        vals = [
            row["order_id"], row["sku"], row["quantity"], row["item_price"], row["payment"],
            eff_cost or 0, total_cost, row["return_reason"], row["product_condition"], row["order_status"],
        ]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(idx, j, v)
            if row.get("is_return"):
                c.fill = orange_fill
    for i in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(report: dict) -> bytes:
    s = report["summary"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)

    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=22, textColor=colors.HexColor("#" + EMERALD), spaceAfter=6)
    subtitle = ParagraphStyle("s", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#5C5F5A"), spaceAfter=18)
    label = ParagraphStyle("l", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#5C5F5A"))

    story = []
    story.append(Paragraph(f"Amazon Monthly P&amp;L — {_month_name(s['target_month'])} {s['target_year']}", title))
    story.append(Paragraph(f"Generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')} • Amazon Seller Reconciliation", subtitle))

    def money(v):
        return f"₹ {v:,.2f}"

    # 3 headline stat cards (label row + value row)
    cards = [
        ["TOTAL RECEIVED", money(s["total_received"])],
        ["TOTAL DEDUCTION", money(s["total_deduction"])],
        ["FINAL PROFIT", money(s["final_profit"])],
    ]
    # Reformat to two rows: label + value
    header_row = [c[0] for c in cards]
    value_row  = [c[1] for c in cards]
    t = Table([header_row, value_row], colWidths=[55*mm]*3)
    t.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 0.5, colors.HexColor("#D8DAD5")),
        ("INNERGRID", (0,0), (-1,-1), 0.5, colors.HexColor("#D8DAD5")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 8),
        ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#5C5F5A")),
        ("FONTNAME", (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE", (0,1), (-1,1), 16),
        ("TEXTCOLOR", (0,1), (-1,1), colors.HexColor("#" + EMERALD)),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    detail_data = [
        ["Settlement",           money(s["settlement"])],
        ["Reimbursement",        money(s["reimbursement"])],
        ["COGS",                 money(s["cogs"])],
        ["Inbound fee",          money(s["inbound_fee"])],
        ["Storage fee",          money(s["storage_fee"])],
        ["Removal fee",          money(s["removal_fee"])],
        ["Ad spend (incl. 18% GST)", money(s["ad_spend"])],
        [f"Packing — Easy Ship ({s.get('easyship_orders_count', 0)} orders)", money(s.get("packing_cost_easyship", 0))],
        ["Packing — Inbound shipments",   money(s.get("packing_cost_inbound", 0))],
        ["Miscellaneous cost",            money(s.get("misc_cost", 0))],
        ["Total Item Price",     money(s["total_item_price"])],
        ["ACOS %",               f"{s['acos_pct']:.2f}%"],
        ["Profit % (revenue)",   f"{s['profit_pct']:.2f}%"],
        ["Profit % on Cost",     f"{s['profit_pct_on_cogs']:.2f}%"],
        ["Customer Return %",    f"{s.get('customer_return_pct', s.get('return_pct', 0)):.2f}%"],
        ["RTO %",                f"{s.get('rto_pct', 0):.2f}%"],
        ["Orders",               str(s["orders_count"])],
        ["Returned Orders",      str(s["returns_count"])],
    ]
    dt = Table([["Metric", "Value"]] + detail_data, colWidths=[110*mm, 55*mm])
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#" + EMERALD)),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#D8DAD5")),
        ("ALIGN", (1,1), (1,-1), "RIGHT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(dt)

    story.append(Spacer(1, 14))
    note_style = ParagraphStyle("gst", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#" + EMERALD), leading=14, spaceBefore=6)
    note_sub   = ParagraphStyle("gsts", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#5C5F5A"), leading=13)
    story.append(Paragraph("<b>Note:</b> Net profit = Gross profit (Final Profit above) − Net GST payable", note_style))
    story.append(Paragraph("GST liability from GSTR filings is not subtracted in this sheet — please deduct your net GST payable separately to arrive at true net profit.", note_sub))

    doc.build(story)
    return buf.getvalue()
